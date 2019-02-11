from tkinter import *
from tkinter import messagebox
from openpyxl import *
import random
import string
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from openpyxl.utils.exceptions import *


# GUI SECTION:
# Create Gui with sizing and title
root = Tk()

root.title("Warranty Form")
root.geometry("600x650")
root.resizable(0, 0)

# Labels for each field
# Width and anchor are used to justify left on single line strings
Rand = Label(root, text="Random Identifier", width=15, anchor="w")
Mn = Label(root, text="Model Number", width=15, anchor="w")
EquipID = Label(root, text="Equipment ID", width=15, anchor="w")
Serial = Label(root, text="Serial Number", width=15, anchor="w")
DofS = Label(root, text="Date Of Service", width=15, anchor="w")
Pn = Label(root, text="Part Number", width=15, anchor="w")
Description = Label(root, text="Description", width=15, anchor="w")
BWMet = Label(root, text="Total Meter", width=15, anchor="w")
ColorMet = Label(root, text="B/W Meter", width=15, anchor="w")
Color = Label(root, text="Color", width=15, anchor="w")
OSN = Label(root, text="Old Serial Number", width=15, anchor="w")
NSN = Label(root, text="New Serial Number", width=15, anchor="w")
DSC = Label(root, text="Problem Description", width=15, anchor="w")

# Place Labels in grid add some padding
Rand.grid(row=0, column=0, padx=10, pady=(40, 10))
Mn.grid(row=1, column=0, padx=10, pady=10)
EquipID.grid(row=2, column=0, padx=10, pady=10)
Serial.grid(row=3, column=0, padx=10, pady=10)
DofS.grid(row=4, column=0, padx=10, pady=10)
Pn.grid(row=5, column=0, padx=10, pady=10)
Description.grid(row=6, column=0, padx=10, pady=10)
BWMet.grid(row=7,column=0, padx=10, pady=10)
ColorMet.grid(row=8, column=0, padx=10, pady=10)
Color.grid(row=9, column=0, padx=10, pady=10)
OSN.grid(row=10, column=0, padx=10, pady=10)
NSN.grid(row=11, column=0, padx=10, pady=10)
DSC.grid(row=12, column=0, padx=10, pady=10)


# MAIN LOGIC SECTION
# This function provides a unique 4 letter/number string to label the parts with
def identifier(size=4, chars=string.ascii_uppercase + string.digits):
    rand_ent_value = (''.join(random.choice(chars) for _ in range(size)))
    return rand_ent_value


# This sets the identifier to the random value coming from the identifier function
ident = StringVar()
ident.set(identifier())

# Init Entry fields
Rand_ent = Entry(root, textvariable=ident, state='disabled')
Mn_ent = Entry(root)
EquipID_ent = Entry(root)
Serial_ent = Entry(root)
DofS_ent = Entry(root)
Pn_ent = Entry(root)
Description_ent = Entry(root)
BWmet_ent = Entry(root)
ColorMet_ent = Entry(root)
Color_ent = Entry(root)
OSN_ent = Entry(root)
NSN_ent = Entry(root)
DSC_ent = Entry(root, width=60)

# Place entry field in respect to labels
Rand_ent.grid(row=0, column=1, padx=10, pady=(40, 10))
Mn_ent.grid(row=1, column=1, padx=10, pady=10)
EquipID_ent.grid(row=2, column=1, padx=10, pady=10)
Serial_ent.grid(row=3, column=1, padx=10, pady=10)
DofS_ent.grid(row=4, column=1, padx=10, pady=10)
Pn_ent.grid(row=5, column=1, padx=10, pady=10)
Description_ent.grid(row=6, column=1, padx=10, pady=10)
BWmet_ent.grid(row=7, column=1, padx=10, pady=10)
ColorMet_ent.grid(row=8, column=1, padx=10, pady=10)
Color_ent.grid(row=9, column=1, padx=10, pady=10)
OSN_ent.grid(row=10, column=1, padx=10, pady=10)
NSN_ent.grid(row=11, column=1, padx=10, pady=10)
DSC_ent.grid(row=12, column=1, columnspan=2, padx=10, pady=10)

# Labels to guide users
Randexp = Label(root, text="Write this on the box for Kevin", width=30, anchor="w")
Randexp.grid(row=0, column=2, padx=10, pady=(50, 10))
Blank = Label(root, text="If N/A leave blank", width=30, anchor="w")
Blank.grid(row=9, column=2, padx=10, pady=10)
Blank1 = Label(root, text="If N/A leave blank", width=30, anchor="w")
Blank1.grid(row=10, column=2, padx=10, pady=10)
Blank2 = Label(root, text="If N/A leave blank", width=30, anchor="w")
Blank2.grid(row=11, column=2, padx=10, pady=10)


# this function will clear all text in the entry boxes
def clear():
    Mn_ent.delete(0, END)
    EquipID_ent.delete(0, END)
    Serial_ent.delete(0, END)
    DofS_ent.delete(0, END)
    Pn_ent.delete(0, END)
    Description_ent.delete(0, END)
    Color_ent.delete(0, END)
    OSN_ent.delete(0, END)
    NSN_ent.delete(0, END)
    DSC_ent.delete(0, END)
    BWmet_ent.delete(0, END)
    ColorMet_ent.delete(0, END)


# this function will clear all fields related to the item for multi part warranties
def anotheritem():
    Pn_ent.delete(0, END)
    Description_ent.delete(0, END)
    Color_ent.delete(0, END)
    OSN_ent.delete(0, END)
    NSN_ent.delete(0, END)


# this initializes the workbook
wb = Workbook()
ws = wb.active


# EMAIL LOGIC SECTION
# this function is used to email the completed warranty to Kevin using TLS for security
def sendit():
    filepathname = Rand_ent.get() + " " + Serial_ent.get() + " " \
            + Mn_ent.get() + " " + DofS_ent.get().replace("/", "-") + ".xlsx"
    filename = os.path.join("Completed Warranties", filepathname)
    # sender credentials are added here removed for privacy.
    gmail_login = ""
    gmail_password = ""
    sender = gmail_login

# Recipient address added here, removed for privacy
    reciever = ""

# SMTPLIB prefers to have toaddr feed a variable
    toaddr = reciever
    cc = []
    bcc = []
    message_subject = filepathname.replace(".xlsx", "")
    message_text = "This is a warranty for " + Serial_ent.get()
    msg = MIMEMultipart()

# EMAIL CONTENTS
    msg["From"] = sender
    msg["To"] = toaddr
    msg["Subject"] = message_subject

# FILE ATTACHMENT
    msg.attach(MIMEText(message_text, "plain"))
    attachment = open(filename, "rb")
    attment = MIMEBase('application', 'octet-stream')
    attment.set_payload(attachment.read())
    encoders.encode_base64(attment)
    attment.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    msg.attach(attment)

# MESSAGE FORMATTING WITH ATTACHMENT
    message = msg.as_string()

    toaddrs = [toaddr] + cc + bcc

# Opens TLS com with gmail server, authenticates, and sends message with attachment
    try:
        s = smtplib.SMTP("smtp.gmail.com", "587")
        s.starttls()
        s.ehlo()
        s.login(gmail_login, gmail_password)
        s.sendmail(from_addr=sender, to_addrs=toaddrs, msg=message)

    except smtplib.SMTPException:
        messagebox.showerror("Error", "Failed to send email")


# SUBMISSION LOGIC SECTION
# this function is used by the new warranty button to clear out the identifier and replace it with a new one
def newident():
    Rand_ent.delete(0, END)
    identifier()
    ident.set(identifier())


# upon hitting submit this pop up will see if the user wants to add a part, start a new warranty, or close the app
def submitpop():
    pop = Tk()
    pop.title("Submitted")
    pop.geometry("420x50")
    filename = Rand_ent.get() + " " + Serial_ent.get() + " " \
        + Mn_ent.get() + " " + DofS_ent.get().replace("/" , "-") + ".xlsx"

# this function cleans up the code a bit as all three buttons will use this code
    def booksaver():
        try:
            wb.save(os.path.join("Completed Warranties", filename))
        except OSError:
            messagebox.showerror("Error", "Please create Completed Warranties Folder")
        except IllegalCharacterError:
            messagebox.showerror("Error", "Invalid Data")
        except ReadOnlyWorkbookException:
            messagebox.showerror("Error", "Read Only WorkBook")

# same as above, cleans code, new warranty and done both use this function
    def booksender():
        try:
            sendit()
        except NameError:
            messagebox.showerror("Error", "Failed to create email")

# logic for adding multi-part warranties
    def popkilladd():
        booksaver()
        anotheritem()
        pop.destroy()

# logic for SAVE > SEND > START OVER
    def popkillnew():
        booksaver()
        booksender()
        clear()
        newident()
        wb.close()
        pop.destroy()

# logic for SAVE > SEND > CLOSE APPLICATION
    def popkilldone():
        booksaver()
        booksender()
        root.destroy()
        pop.destroy()

# These are used to add more items, submit current warranty and start a new one, or submit current warranty and finish
    addbutton = Button(pop, text="MORE ITEMS", width=15, anchor="center", padx=10, pady=10, command=popkilladd)
    addbutton.grid(row=1, column=1)
    clearbutton = Button(pop, text="NEW WARRANTY", width=15, anchor="center", padx=10, pady=10, command=popkillnew)
    clearbutton.grid(row=1, column=2)
    donebutton = Button(pop, text="DONE", width=15, anchor="center", padx=10, pady=10, command=popkilldone)
    donebutton.grid(row=1, column=3)


# This is used instead of message box as it allows unified control of error text
def fillpls(section):
    fill = Tk()
    fill.title("Attention")
    fill.geometry("340x200")

# tkinter prefers to call a function rather than execute a built in so this function is used
    def fillkill():
        fill.destroy()

    checksection = Label(fill, text="Please enter " + section, width=40, anchor="center")
    checksection.grid(row=0, column=0, padx=25, pady=35)

    okbutton = Button(fill, text="OK", width=15, anchor="center", padx=10, pady=10, command=fillkill)
    okbutton.grid(row=1, column=0, padx=20, pady=20)


# this function will check user inputs to prevent forgotten info
def inputcheck(entrystr):
    if entrystr == "":
        return TRUE
    elif entrystr.isspace():
        return TRUE
    elif len(entrystr) < 2:
        return TRUE
    elif len(entrystr) > 230:
        return TRUE
    else:
        return FALSE


# this function is where the magic happens, it checks inputs then politely shoves them into an excel document
def submit():
    if inputcheck(Mn_ent.get()):
        fillpls("Model Number")
    elif inputcheck(EquipID_ent.get()):
        fillpls("ID Number")
    elif inputcheck(Serial_ent.get()):
        fillpls("Serial Number")
    elif inputcheck(DofS_ent.get()):
        fillpls("Date Of Service")
    elif inputcheck(Description_ent.get()):
        fillpls("Part Description")
    elif BWmet_ent.get().islower() or BWmet_ent.get().isupper() or inputcheck(BWmet_ent.get()):
        fillpls("numbers only for Total Meter")
    elif ColorMet_ent.get().islower() or ColorMet_ent.get().isupper() or inputcheck(ColorMet_ent.get()):
        fillpls("numbers only for B/W Meter")
    elif inputcheck(DSC_ent.get()):
        fillpls("less than 240 characters\n for your problem description.")
    else:
        # setting variables for current row and column
        current_row = ws.max_row
        ws.cell(row=current_row + 1, column=1).value = Rand_ent.get()
        ws.cell(row=current_row + 1, column=2).value = Mn_ent.get()
        ws.cell(row=current_row + 1, column=3).value = EquipID_ent.get()
        ws.cell(row=current_row + 1, column=4).value = Serial_ent.get()
        ws.cell(row=current_row + 1, column=5).value = DofS_ent.get()
        ws.cell(row=current_row + 1, column=6).value = Pn_ent.get()
        ws.cell(row=current_row + 1, column=7).value = Description_ent.get()
        ws.cell(row=current_row + 1, column=8).value = BWmet_ent.get()
        ws.cell(row=current_row + 1, column=9).value = str(int(BWmet_ent.get()) - int(ColorMet_ent.get()))

        # these if/else blocks will input N/A for color and serial numbers if they do not apply to the warranty
        if inputcheck(Color_ent.get()):
            ws.cell(row=current_row + 1, column=10).value = "N/A"
        else:
            ws.cell(row=current_row + 1, column=10).value = Color_ent.get()
        if inputcheck(OSN_ent.get()):
            ws.cell(row=current_row + 1, column=11).value = "N/A"
        else:
            ws.cell(row=current_row + 1, column=11).value = OSN_ent.get()
        if inputcheck(NSN_ent.get()):
            ws.cell(row=current_row + 1, column=12).value = "N/A"
        else:
            ws.cell(row=current_row + 1, column=12).value = NSN_ent.get()

        ws.cell(row=current_row + 1, column=13).value = DSC_ent.get()
        submitpop()


# SpreadSheet piping
def excelwrite():
    # This is to resize the columns for ease of reading
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 40

    # This will write the appropriate headers for the spreadsheet
    ws.cell(row=1, column=1).value = "Identifier"
    ws.cell(row=1, column=2).value = "Model Number"
    ws.cell(row=1, column=3).value = "Equipment ID"
    ws.cell(row=1, column=4).value = "Machine Serial Number"
    ws.cell(row=1, column=5).value = "Date Of Service"
    ws.cell(row=1, column=6).value = "Item Part Number"
    ws.cell(row=1, column=7).value = "Item Description"
    ws.cell(row=1, column=8).value = "Total Meter"
    ws.cell(row=1, column=9).value = "Color Meter"
    ws.cell(row=1, column=10).value = "Color"
    ws.cell(row=1, column=11).value = "Old Serial Number"
    ws.cell(row=1, column=12).value = "New Serial Number"
    ws.cell(row=1, column=13).value = "Problem Description"


# Submit button
submitButton = Button(root, text="SUBMIT", width=11, anchor="center", command=submit)
submitButton.grid(row=13, column=0, pady=(10, 10))
# Clear Button
clearButton = Button(root, text="CLEAR ALL", width=11, anchor="center", command=clear)
clearButton.grid(row=13, column=1, pady=(10, 10))


# Start Gui and generate excel doc but in the opposite order
excelwrite()
root.mainloop()

